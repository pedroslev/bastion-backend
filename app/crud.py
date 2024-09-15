import os
import random
import requests
from config import IMAGE_IA_DIR, IMAGE_WORDCLOUD_DIR, SERVE_IMAGE, EXCEL_DIR, MOSAIC_DIR
from dotenv import load_dotenv
import logging
from openpyxl import Workbook, load_workbook
from datetime import datetime
import math
from PIL import Image

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def save_text(text: str):
    try:
        logger.info(f"Saving text: {text}")
        text_file = open(f"{IMAGE_WORDCLOUD_DIR}/palabras.txt", "a")
        text_file.write(text + "\n")
        text_file.close()
        return True
    except Exception as e:
        logger.error(f"Error saving text: {str(e)}") 
        return False

def get_all_texts():
    try:
        text_file = open(f"{IMAGE_WORDCLOUD_DIR}/palabras.txt", "r")
        texts = text_file.readlines()
        text_list = [text.strip() for text in texts]
        text_file.close()
        return text_list
    except Exception as e:
        logger.error(f"Error getting all texts: {str(e)}")
        return False

# Guardar la imagen en el directorio local especificado en el archivo .env
def save_image_locally(image_url: str, image_name: str):
    try:
        response = requests.get(image_url)
        image_content = response.content
        save_dir = IMAGE_IA_DIR + image_name
        with open(save_dir, "wb") as file:
            file.write(image_content)
        logger.info(f"Image saved locally at: {save_dir}")
        return save_dir
    except Exception as e:
        logger.error(f"Error saving image locally: {str(e)}")
        return False

def get_image_filenames():
    try:
        # Recuperar nombres de imágenes del directorio local
        image_dir = os.path.dirname(IMAGE_IA_DIR)
        image_filenames = os.listdir(image_dir)
        #add to each image the url to access it
        image_filenames = [f"{SERVE_IMAGE}/files/ia/{image}" for image in image_filenames]
        return image_filenames
    except Exception as e:
        logger.error(f"Error getting image filenames: {str(e)}")
        return False

def get_word_cloud_image():
    try:
    # Recuperar nombres de imágenes del directorio local
        image_dir = os.path.dirname(IMAGE_WORDCLOUD_DIR)
        image_filenames = os.listdir(image_dir)
        #add to each image the url to access it
        image_filenames = [f"{SERVE_IMAGE}/files/wordcloud/{image}" for image in image_filenames]
        #only return the one with the name wordcloud.png
        image_filenames = [image for image in image_filenames if "wordcloud.png" in image]
        return image_filenames[0]
    except Exception as e:
        logger.error(f"Error getting image filenames: {str(e)}")
        return False
    
def insert_text_to_excel(filename, text):
    # Get current date and timestamp
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    file_path = os.path.join(EXCEL_DIR, filename)
    # Check if file exists
    if os.path.exists(file_path):
        # Load existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Horario", "Respuesta"])  # Add header if new file

    # Append new row with date, timestamp, and text
    sheet.append([current_date, current_time, text])

    # Save the workbook
    workbook.save(file_path)

def insert_momento4_to_excel(filename, data):
    # Get current date and timestamp
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    file_path = os.path.join(EXCEL_DIR, filename)
    # Check if file exists
    if os.path.exists(file_path):
        # Load existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Horario", "Summit content", "Day 1 presentations","Speaker day 1", "Speaker day 2", "Location", "Food", "Duration", "Summit Overall", "Liked the most", "Liked the least"])  # Add header if new file

    # Append new row with date, timestamp, and text
    sheet.append([current_date, current_time, data.content,data.presentation1,data.speaker1,data.speaker2,data.location,data.food,data.duration,data.summit, data.question1, data.question2])

    # Save the workbook
    workbook.save(file_path)

def insert_training_to_excel(filename, data):
    # Get current date and timestamp
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    file_path = os.path.join(EXCEL_DIR, filename)
    # Check if file exists
    if os.path.exists(file_path):
        # Load existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Horario", "Awards Event", "Wednesday training","Thursday Training", "Likes", "Dislikes"])  # Add header if new file

    # Append new row with date, timestamp, and text
    sheet.append([current_date, current_time, data.AwardsEvent, data.TrainingWednesday, data.TrainingThursday, data.question1, data.question2])

    # Save the workbook
    workbook.save(file_path)

def insert_ia_to_excel(filename, text, name, image):
    # Get current date and timestamp
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    file_path = os.path.join(EXCEL_DIR, filename)
    # Check if file exists
    if os.path.exists(file_path):
        # Load existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Horario", "Nombre", "Respuesta", "Imagen"])

    # Append new row with date, timestamp, and text
    sheet.append([current_date, current_time,name, text, image])

    # Save the workbook
    workbook.save(file_path)


def insert_text_to_excel_momento(filename, data):
    # Get current date and timestamp
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    file_path = os.path.join(EXCEL_DIR, filename)
    respuesta1 = data.pregunta1
    respuesta2 = data.pregunta2
    respuesta3 = data.pregunta3
    logger.info(f"Inserting momento data to Excel: {respuesta1}, {respuesta2}, {respuesta3}")

    # Check if file exists
    if os.path.exists(file_path):
        # Load existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Horario", "Respuesta1","Respuesta2", "Respuesta3" ])  # Add header if new file

    # Append new row with date, timestamp, and text
    sheet.append([current_date, current_time, respuesta1, respuesta2, respuesta3])

    # Save the workbook
    workbook.save(file_path)


def get_momento():
    file_path = os.path.join(EXCEL_DIR, "latamTopInitiatives.xlsx")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File {file_path} does not exist.")

    # Load the Excel workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Read the data into lists for each column (excluding the header row)
    respuesta1_list = []
    respuesta2_list = []
    respuesta3_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # min_row=2 skips the header
        respuesta1_list.append(row[2])  # Column 3: Respuesta1
        respuesta2_list.append(row[3])  # Column 4: Respuesta2
        respuesta3_list.append(row[4])  # Column 5: Respuesta3

    # Select 3 random answers from each list
    respuesta1 = random.sample(respuesta1_list, min(3, len(respuesta1_list)))
    respuesta2 = random.sample(respuesta2_list, min(3, len(respuesta2_list)))
    respuesta3 = random.sample(respuesta3_list, min(3, len(respuesta3_list)))

    # Return the selected answers in the required format
    return {
        "respuesta1": respuesta1,
        "respuesta2": respuesta2,
        "respuesta3": respuesta3
    }

def get_momento2():
    file_path = os.path.join(EXCEL_DIR, "biggestTakeaway.xlsx")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File {file_path} does not exist.")

    # Load the Excel workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Read the data into lists for each column (excluding the header row)
    respuesta_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # min_row=2 skips the header
        respuesta_list.append(row[2])  # Column 3: Respuesta1

    # Select 3 random answers from each list
    respuesta1 = random.sample(respuesta_list, min(4, len(respuesta_list)))


    # Return the selected answers in the required format
    return {
        "respuestas": respuesta1,
    }

def create_ia_mosaic(mosaic_width=1000, mosaic_height=1000, image_size=(100, 100), output_file="mosaic.png"):
    # Get all .png files from the directory
    images = [os.path.join(IMAGE_IA_DIR, f) for f in os.listdir(IMAGE_IA_DIR) if f.endswith(".png")]
    logger.info(f"Found {len(images)} images in the directory: {IMAGE_IA_DIR}")
    
    if not images:
        raise ValueError(f"No PNG files found in the directory: {IMAGE_IA_DIR}")
    
    # Determine number of images
    num_images = len(images)
    
    # Calculate the aspect ratio of the mosaic canvas
    aspect_ratio = mosaic_width / mosaic_height
    
    # Estimate number of columns based on the aspect ratio and the number of images
    num_columns = math.ceil(math.sqrt(num_images * aspect_ratio))
    num_rows = math.ceil(num_images / num_columns)
    
    # Calculate the size of each image so they fit perfectly into the mosaic
    image_width = mosaic_width // num_columns
    image_height = mosaic_height // num_rows
    
    logger.info(f"Creating a mosaic with {num_rows} rows and {num_columns} columns.")
    logger.info(f"Each image will be resized to {image_width}x{image_height} pixels.")
    
    # Create a blank canvas for the collage
    collage = Image.new('RGB', (mosaic_width, mosaic_height), color='white')

    # Resize and paste each image into the collage grid
    for index, image_path in enumerate(images):
        row = index // num_columns
        col = index % num_columns
        
        with Image.open(image_path) as img:
            img = img.resize((image_width, image_height), Image.Resampling.LANCZOS)
            collage.paste(img, (col * image_width, row * image_height))
    
    # Save the final collage to a file
    output_file = os.path.join(MOSAIC_DIR, output_file)
    collage.save(output_file)
    logger.info(f"Mosaic saved to {output_file}")
    
    return True